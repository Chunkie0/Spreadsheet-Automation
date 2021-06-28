import openpyxl

inventoryFile = openpyxl.load_workbook("inventory.xlsx")

productList = inventoryFile["Sheet1"]

productsPerSupplier = {}
totalValuePerSupplier = {}
productsInventoryUnder10 = {}

for productRow in range(2, productList.max_row + 1):
    supplierName = productList.cell(productRow, 4).value
    inventory = productList.cell(productRow, 2).value
    value = productList.cell(productRow, 3).value
    inventoryPrice = productList.cell(productRow, 5)

    # dictionary values to products per supplier
    if supplierName in productsPerSupplier:
        currentNumProducts = productsPerSupplier.get(supplierName)
        productsPerSupplier[supplierName] = currentNumProducts + 1
    else:
        productsPerSupplier[supplierName] = 1

    
    # dictionary values for total value per supplier
    if supplierName in totalValuePerSupplier:
        currentValueProducts = totalValuePerSupplier.get(supplierName)
        totalValuePerSupplier[supplierName] = currentValueProducts + inventory * value
    else:
        totalValuePerSupplier[supplierName] = inventory * value

    # dictionary values for inventory less than 10
    if inventory < 10:
        productsInventoryUnder10[productRow-1] = int(inventory)

    # creates new column with inventory prices and saves in a new file
    inventoryPrice.value = inventory * value


inventoryFile.save("new_inventory.xlsx") # saves as a new file with changes

print(totalValuePerSupplier) 
print(productsPerSupplier)
print(productsInventoryUnder10)